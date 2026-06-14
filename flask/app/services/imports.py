from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any

from flask import current_app
from openpyxl import load_workbook

from app.extensions import db
from app.models import BaseNote, ImportJob
from app.services.geocoding import build_geocode_query, geocode_note


EXPECTED_COLUMNS = [
    "Land",
    "Region/BL/Ort",
    "Nummer",
    "Schein_Front_URL",
    "Schein_Front_Datei",
    "Schein_Back_URL",
    "Schein_Back_Datei",
    "Bezeichnung",
    "Jahr",
    "Adresse",
]

OPTIONAL_COLUMNS = [
    "Variante",
]


@dataclass
class ParsedRows:
    rows: list[dict[str, Any]]
    warnings: list[str]


def parse_excel(file_storage) -> ParsedRows:
    workbook = load_workbook(BytesIO(file_storage.read()), data_only=True)
    sheet = workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    warnings: list[str] = []

    missing_columns = [column for column in EXPECTED_COLUMNS if column not in headers]
    if missing_columns:
        warnings.append("Fehlende Spalten: " + ", ".join(missing_columns))

    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw = dict(zip(headers, values))
        row = {column: raw.get(column) for column in EXPECTED_COLUMNS + OPTIONAL_COLUMNS}
        row["_row_number"] = row_number
        rows.append(row)
    return ParsedRows(rows=rows, warnings=warnings)


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_year(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "title": clean_text(row.get("Bezeichnung")),
        "country": clean_text(row.get("Land")),
        "region_or_city": clean_text(row.get("Region/BL/Ort")),
        "catalog_number": clean_text(row.get("Nummer")),
        "front_img": clean_text(row.get("Schein_Front_URL")) or clean_text(row.get("Schein_Front_Datei")),
        "back_img": clean_text(row.get("Schein_Back_URL")) or clean_text(row.get("Schein_Back_Datei")),
        "issue_year": normalize_year(row.get("Jahr")),
        "variant_type": clean_text(row.get("Variante")),
        "address": clean_text(row.get("Adresse")),
        "_row_number": row.get("_row_number"),
    }


def create_import_preview(filename: str, rows: list[dict[str, Any]], warnings: list[str]) -> ImportJob:
    seen: set[str] = set()
    payload: list[dict[str, Any]] = []
    counts = {
        "new_count": 0,
        "update_count": 0,
        "unchanged_count": 0,
        "skipped_count": 0,
        "geocode_count": 0,
    }

    for raw in rows:
        item = normalize_row(raw)
        catalog_number = item["catalog_number"]
        if not item["title"] or not item["country"] or not catalog_number:
            item["_action"] = "skipped"
            item["_reason"] = "Titel, Land oder Nummer fehlt"
            counts["skipped_count"] += 1
        elif catalog_number in seen:
            item["_action"] = "skipped"
            item["_reason"] = f"Doppelte Katalognummer in Excel: {catalog_number}"
            warnings.append(item["_reason"])
            counts["skipped_count"] += 1
        else:
            seen.add(catalog_number)
            existing = BaseNote.query.filter_by(catalog_number=catalog_number).first()
            if existing is None:
                item["_action"] = "new"
                counts["new_count"] += 1
            elif _note_changed(existing, item):
                item["_action"] = "update"
                counts["update_count"] += 1
            else:
                item["_action"] = "unchanged"
                counts["unchanged_count"] += 1

            needs_geocode = bool(
                item.get("country")
                and (item.get("address") or item.get("region_or_city"))
                and (existing is None or not existing.has_coordinates)
            )
            if needs_geocode:
                counts["geocode_count"] += 1
        payload.append(item)

    job = ImportJob(
        filename=filename,
        status="preview",
        total_rows=len(rows),
        warnings=warnings,
        payload=payload,
        **counts,
    )
    db.session.add(job)
    db.session.commit()
    return job


def _note_changed(note: BaseNote, item: dict[str, Any]) -> bool:
    fields = [
        "title",
        "country",
        "region_or_city",
        "address",
        "issue_year",
        "variant_type",
        "front_img",
        "back_img",
    ]
    return any(getattr(note, field) != item.get(field) for field in fields)


def _import_row(item: dict[str, Any]) -> None:
    note = BaseNote.query.filter_by(catalog_number=item["catalog_number"]).first()
    if note is None:
        note = BaseNote(catalog_number=item["catalog_number"], is_active=True)
        db.session.add(note)

    for field in [
        "title",
        "country",
        "region_or_city",
        "address",
        "issue_year",
        "variant_type",
        "front_img",
        "back_img",
    ]:
        setattr(note, field, item.get(field))

    if (
        current_app.config["IMPORT_GEOCODE_DURING_IMPORT"]
        and build_geocode_query(note)
        and not note.has_coordinates
    ):
        geocode_note(note)


def run_import_batch(job: ImportJob, batch_size: int | None = None) -> ImportJob:
    if job.status == "completed":
        return job

    batch_size = batch_size or current_app.config["IMPORT_BATCH_SIZE"]
    payload = job.payload or []
    start = min(job.processed_rows, len(payload))
    end = min(start + batch_size, len(payload))

    job.status = "running"
    db.session.commit()

    for item in payload[start:end]:
        if item.get("_action") == "skipped":
            job.processed_rows += 1
            continue

        _import_row(item)
        job.processed_rows += 1
        db.session.flush()

    if job.processed_rows >= len(payload):
        job.status = "completed"
    db.session.commit()
    return job


def run_import(job: ImportJob) -> ImportJob:
    while job.status != "completed":
        run_import_batch(job, batch_size=current_app.config["IMPORT_BATCH_SIZE"])
    return job
